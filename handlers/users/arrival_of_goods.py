import glob
import os

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from data.config import path


def actual_date():
    import datetime

    # находим текущую дату
    today = datetime.date.today()

    # вычисляем дату минус два дня от текущей
    two_days_from_today = today - datetime.timedelta(days=5)
    compare_date = pd.to_datetime(f'{two_days_from_today}')
    # создаем список дат
    try:
        df1 = pd.read_excel(f'{path}/files/file_arrival/График поставок.xlsx',
                            usecols=['Код графика', 'Тип операции', 'Объем', 'Планируемая дата прихода в магазин'])
        df2 = pd.read_excel(f'{path}/files/file_arrival/График перемещений и мебели.xlsx',
                            usecols=['Код графика', 'Тип операции', 'Объем', 'Планируемая дата прихода в магазин'])
        df_date = pd.concat([df1, df2])
        df_date.to_excel('asdasdas.xlsx')
        df_date = df_date[df_date['Планируемая дата прихода в магазин'] > compare_date]
        list_ds_files = []
        folder_path = f"{path}/files/file_arrival/DSs/"
        files = os.listdir(folder_path)
        for file in files:
            if os.path.isfile(os.path.join(folder_path, file)):
                filename, extension = os.path.splitext(file)
                list_ds_files.append(filename)
        df_date = df_date[df_date['Код графика'].isin(list_ds_files)]
        df_date.to_csv(f'{path}/files/file_arrival/keyboards.csv')
        return df_date
    except:
        logger.debug('Ошибка при чтении файла с дсками')


def open_file_ds():
    list_ds = []
    df = actual_date()

    df1 = pd.read_csv(f'{path}/files/file_011_825.csv')
    df2 = pd.read_csv(f'{path}/files/file_012_825.csv')
    df3 = pd.read_csv(f'{path}/files/file_S_825.csv')
    df4 = pd.read_csv(f'{path}/files/file_V_Sales.csv')
    result = pd.concat([df1, df2, df3, df4])
    result = result[~(result['Доступно'].isnull())]
    result = result.rename(columns={'Код \nноменклатуры': 'Номенклатура'})
    result = result.groupby(['Номенклатура']).agg(
        {'Номенклатура': 'first', 'Описание товара': 'first', 'ТГ': 'first', 'Доступно': 'sum'})
    result.reset_index(drop=True, inplace=True)

    if not df.empty:
        df['union'] = df['Код графика'] + ' ' + df['Планируемая дата прихода в магазин'].apply(
            lambda x: pd.to_datetime(x).strftime("%d.%m.%Y")) + ' ' + df['Тип операции']+ \
                           ' ' + df['Объем'].apply(lambda x: str(x))
        list_ds = df['union'].to_list()
    if list_ds:
        for ds in list_ds:
            try:
                name = ds.split()[0]

                merged_df = out_df_merged(name)

                grouped_df = out_df_grouped(name)

                new_products(union_df=result, name=name)


            except Exception as ex:
                logger.debug(f'{ds} - {ex}')


def out_df_merged(name):
    if os.path.exists(f'{path}/files/file_arrival/result/{name}.xlsx'):
        merged_df = pd.read_excel(f'{path}/files/file_arrival/result/{name}.xlsx')
    else:
        df_ds = pd.read_excel(f'{path}/files/file_arrival/DSs/{name}.xlsx',
                              dtype={'Код номенклатуры': str, 'Наименование номенклатуры': str,
                                     'Объем': float})
        try:
            df_ds = df_ds.rename(columns={'Отгруженное количество': 'Количество'})
        except Exception:
            ...
        df_ds = df_ds.rename(columns={'Код номенклатуры': 'Номенклатура'})
        df_min = pd.read_excel(f'{path}/files/file_arrival/art_tg.xlsx',
                               dtype={'Номенклатура': str, 'name': str, 'SG': str})
        merged_df = pd.merge(df_ds, df_min[['Номенклатура', 'SG']], on='Номенклатура', how='left')
        merged_df = merged_df.groupby(['Номенклатура']).agg({'Номенклатура': 'first',
                                                             'Наименование номенклатуры': 'first',
                                                             'Количество': 'sum',
                                                             'Объем': 'sum',
                                                             'SG': 'first',
                                                             })

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(merged_df, index=False, header=True):
        ws.append(r)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    wb.save(f'{path}/files/file_arrival/result/{name}.xlsx')
    return merged_df


def out_df_grouped(name):
    if os.path.exists(f'{path}/files/file_arrival/result/{name}_grouped.xlsx'):
        grouped_df = pd.read_excel(f'{path}/files/file_arrival/result/{name}_grouped.xlsx')
    else:
        merged_df = out_df_merged(name)
        merged_df['SG'] = merged_df['SG'].fillna('Нет данных о ТГ')
        merged_df = merged_df.assign(count=1)
        try:
            grouped_df = merged_df.groupby(['SG']).agg(
                {'SG': 'first', 'Количество': 'sum', 'Объем': 'sum', 'count': 'count'
                 })
        except Exception:
            grouped_df = merged_df.groupby(['SG']).agg(
                {'SG': 'first', 'Количество': 'sum', 'Объем': 'sum', 'count': 'count'
                 })
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(grouped_df, index=False, header=True):
            ws.append(r)
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        wb.save(f'{path}/files/file_arrival/result/{name}_grouped.xlsx')

    return grouped_df


def new_products(union_df, name):
    df_merged = out_df_merged(name)
    result_new = pd.merge(df_merged, union_df, on='Номенклатура', how='left')
    result_new = result_new[result_new['Доступно'].isnull()]

    result_new = result_new.drop(['Описание товара', 'ТГ', 'Доступно'], axis=1)
    result_new = result_new.rename(columns={'SG': 'ТГ'})

    try:
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(result_new, index=False, header=True):
            try:
                ws.append(r)
            except Exception as ex:
                logger.debug(ex)
        for column_cells in ws.columns:
            try:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
            except Exception as ex:
                logger.debug(ex)

        wb.save(f'{path}/files/file_arrival/result/{name}_result_new.xlsx')
    except Exception as ex:
        logger.debug(ex)


if __name__ == '__main__':
    open_file_ds()
